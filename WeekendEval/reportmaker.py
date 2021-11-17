import pathlib
import logging
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

# takes in a month name as well as the file name
# Parses the input based on the () symbols to seperate the two
def load_report(input):

    #parse the input based on the () symbols and use what is inbetween as the filepath
    fp = input[input.find("(")+1:input.find(")")]

    #parse the input and take what is before the ( symbol as the month name
    month = input.partition('(')[0]

    workbook = load_workbook(filename=fp)
    #log if the sheet name from the input is not found


    # Convert the summary rolling sheet into a dictionary with the key being month name
    sheet = workbook['Summary Rolling MoM']
    summary_ss_data = {}

    rows = sheet.iter_rows(min_row = 2, max_row = 13, min_col = 1, max_col = 6)
    for row in rows:
        row_data = {}
        row_name = str(row[0].value.strftime("%B"))
        row_data['Calls Offered'] = row[1].value
        row_data['Abandon after 30s'] = row[2].value
        row_data['FCR'] = row[3].value
        row_data['DSAT'] = row[4].value
        row_data['CSAT'] = row[5].value

        #create a dictionary entry with key value as the month name and the value as a list of all the values
        summary_ss_data[row_name] = row_data
    
    print(summary_ss_data)

    # Logging the report based off the month given and taking the values from the dictionary
    logging.basicConfig(filename='expedia_report.log', encoding='utf-8', level=logging.DEBUG)
    logging.info('Summary for ' + month )
    logging.info('Calls Offered: ' + str(summary_ss_data[month]['Calls Offered']))
    logging.info('Abandon After 30s: ' + str( (summary_ss_data[month]['Abandon after 30s']) * 100)  + '%')
    logging.info('FCR : ' + str( (summary_ss_data[month]['FCR'])* 100)  + '%')
    logging.info('DSAT : ' + str( (summary_ss_data[month]['DSAT'])* 100) + '%')
    logging.info('CSAT : ' + str( (summary_ss_data[month]['CSAT'])* 100) + '%')


def voc_report(fp):
    workbook = load_workbook(filename=fp)
    #convert the VOC Rolling MoM sheet into a dictionary with the key being the month name
    sheet2 = workbook['VOC Rolling MoM']
    voc_ss_data = {}
    col = sheet2['C']
    
    #create an empty dict to hold the values
    col_data = {}
    col_data['Promoters'] = col[3].value
    col_data['Passives'] = col[5].value
    col_data['Detractors'] = col[7].value

    voc_ss_data[str(col[0].value.strftime("%B"))] = col_data

    logging.basicConfig(filename='expedia_report.log', encoding='utf-8', level=logging.DEBUG)
    logging.info('Summary for VOC Rolling MoM of' + str(col[0].value))

    #If-Elseif statements based on the promoter score. Log either good or bad
    if (voc_ss_data['January']['Promoters'] > 200):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Promoters']) + " : Good" )
    elif (voc_ss_data['January']['Promoters'] <= 200):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Promoters']) + " : Bad" )

    #If-Elseif statements based on the Passives score. Log either good or bad
    if (voc_ss_data['January']['Passives'] > 100):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Passives']) + " : Good" )
    elif (voc_ss_data['January']['Passives'] <= 100):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Passives']) + " : Bad" )

    #If-Elseif statements based on the Detractors score. Log either good or bad
    if (voc_ss_data['January']['Detractors'] < 100):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Detractors']) + " : Good" )
    elif (voc_ss_data['January']['Detractors'] >= 100):
        logging.info('Promoter score of ' + str(voc_ss_data['January']['Detractors']) + " : Bad" )



# voc_report('Python-Basics\WeekendEval\expedia_report_monthly_january_2018.xlsx')
# load_report("March(Python-Basics\WeekendEval\expedia_report_monthly_january_2018.xlsx)")



